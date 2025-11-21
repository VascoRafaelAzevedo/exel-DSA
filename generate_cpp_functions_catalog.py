"""
Comprehensive C++ DSA Functions and Algorithms Catalog Generator
Generates a multi-sheet Excel workbook with detailed information about:
- C++ STL algorithms (50+ most important algorithms from <algorithm> and <numeric>)
- Container member functions (vector, map, set, list, deque, stack, queue, etc.)
- Usage frequencies for real-world and competitive programming/DSA

Run: python generate_cpp_functions_catalog.py
Requires: pip install pandas openpyxl
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 80)
print("C++ DSA FUNCTIONS & ALGORITHMS CATALOG GENERATOR")
print("=" * 80)
print()


# ==============================================================================
# HELPER FUNCTION FOR FORMATTING
# ==============================================================================

# Constants for cell formatting
UNFILLED_COLOR = '00000000'  # Default unfilled cell color in openpyxl

def format_worksheet(ws, header_color='4472C4', freeze_panes=True):
    """Apply professional formatting to worksheet"""
    # Header styling
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_side = Side(border_style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Apply to header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Freeze panes (header row and first column)
    if freeze_panes:
        ws.freeze_panes = 'B2'
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 3, 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Alternate row colors
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row_idx % 2 == 0:
            for cell in row:
                if cell.fill.start_color.index == UNFILLED_COLOR:  # Only if not already filled
                    cell.fill = light_fill
        for cell in row:
            cell.border = border


# ==============================================================================
# STL ALGORITHMS - Complete <algorithm> header coverage
# ==============================================================================

stl_algorithms = [
    # ========== NON-MODIFYING SEQUENCE OPERATIONS ==========
    {
        "Function": "std::all_of", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, UnaryPredicate p",
        "Arg_Explanation": "first, last: input range; p: unary predicate function",
        "Return_Type": "bool",
        "Description": "Checks if all elements satisfy predicate",
        "When_To_Use": "Validate all elements meet condition; early termination on first false",
        "When_NOT_To_Use": "Need count of matches; checking empty ranges",
        "Real_World_Freq": 5, "DSA_LeetCode_Freq": 8,
        "Example": "bool all_pos = std::all_of(v.begin(), v.end(), [](int x){ return x > 0; });",
        "Notes": "Returns true for empty range; short-circuits",
        "Since_Version": "C++11", "Related": "any_of, none_of"
    },
    {
        "Function": "std::any_of", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, UnaryPredicate p",
        "Arg_Explanation": "first, last: input range; p: unary predicate function",
        "Return_Type": "bool",
        "Description": "Checks if at least one element satisfies predicate",
        "When_To_Use": "Existence check; early termination on first match",
        "When_NOT_To_Use": "Need all matches; get iterator to match (use find_if)",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 8,
        "Example": "bool has_even = std::any_of(v.begin(), v.end(), [](int x){ return x % 2 == 0; });",
        "Notes": "Returns false for empty range; short-circuits",
        "Since_Version": "C++11", "Related": "all_of, none_of, find_if"
    },
    {
        "Function": "std::none_of", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, UnaryPredicate p",
        "Arg_Explanation": "first, last: input range; p: unary predicate function",
        "Return_Type": "bool",
        "Description": "Checks if no elements satisfy predicate",
        "When_To_Use": "Verify condition never occurs; validation",
        "When_NOT_To_Use": "Need to know which elements match",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 6,
        "Example": "bool no_neg = std::none_of(v.begin(), v.end(), [](int x){ return x < 0; });",
        "Notes": "Equivalent to !any_of; returns true for empty range",
        "Since_Version": "C++11", "Related": "all_of, any_of"
    },
    {
        "Function": "std::for_each", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, UnaryFunction f",
        "Arg_Explanation": "first, last: input range; f: function to apply to each element",
        "Return_Type": "UnaryFunction",
        "Description": "Applies function to each element in range",
        "When_To_Use": "Side effects on each element; custom iteration logic",
        "When_NOT_To_Use": "Modern C++ prefers range-for; pure transformation (use transform)",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 4,
        "Example": "std::for_each(v.begin(), v.end(), [](int& x){ std::cout << x << ' '; });",
        "Notes": "Can modify if function takes reference parameter",
        "Since_Version": "C++98", "Related": "range-for, transform"
    },
    {
        "Function": "std::for_each_n", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, Size n, UnaryFunction f",
        "Arg_Explanation": "first: start iterator; n: number of elements; f: function to apply",
        "Return_Type": "InputIt",
        "Description": "Applies function to first n elements",
        "When_To_Use": "Process specific number of elements; early termination",
        "When_NOT_To_Use": "Process entire range (use for_each); conditional (use find_if)",
        "Real_World_Freq": 3, "DSA_LeetCode_Freq": 2,
        "Example": "std::for_each_n(v.begin(), 5, [](int x){ std::cout << x; });",
        "Notes": "Returns iterator past last processed element",
        "Since_Version": "C++17", "Related": "for_each"
    },
    {
        "Function": "std::count", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, const T& value",
        "Arg_Explanation": "first, last: input range; value: value to count",
        "Return_Type": "iterator_traits<InputIt>::difference_type",
        "Description": "Counts occurrences of value in range",
        "When_To_Use": "Simple value frequency; specific element count",
        "When_NOT_To_Use": "All frequencies (use map); conditional count (use count_if)",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 7,
        "Example": "int cnt = std::count(v.begin(), v.end(), 42);",
        "Notes": "Linear complexity; sorted ranges can use binary search",
        "Since_Version": "C++98", "Related": "count_if, find"
    },
    {
        "Function": "std::count_if", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, UnaryPredicate p",
        "Arg_Explanation": "first, last: input range; p: predicate function",
        "Return_Type": "iterator_traits<InputIt>::difference_type",
        "Description": "Counts elements satisfying predicate",
        "When_To_Use": "Conditional counting; complex filtering criteria",
        "When_NOT_To_Use": "Simple value count (use count); all elements",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 9,
        "Example": "int pos = std::count_if(v.begin(), v.end(), [](int x){ return x > 0; });",
        "Notes": "Very common in DSA; works well with lambdas",
        "Since_Version": "C++98", "Related": "count, find_if"
    },
    {
        "Function": "std::mismatch", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt1 first1, InputIt1 last1, InputIt2 first2",
        "Arg_Explanation": "first1, last1: first range; first2: start of second range",
        "Return_Type": "pair<InputIt1, InputIt2>",
        "Description": "Finds first position where two ranges differ",
        "When_To_Use": "Compare sequences; find divergence point",
        "When_NOT_To_Use": "Full equality check (use equal); single value search",
        "Real_World_Freq": 3, "DSA_LeetCode_Freq": 4,
        "Example": "auto [it1, it2] = std::mismatch(v1.begin(), v1.end(), v2.begin());",
        "Notes": "Returns iterators to first mismatch or end",
        "Since_Version": "C++98", "Related": "equal, find"
    },
    {
        "Function": "std::find", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, const T& value",
        "Arg_Explanation": "first, last: input range; value: value to find",
        "Return_Type": "InputIt",
        "Description": "Finds first occurrence of value",
        "When_To_Use": "Linear search; unsorted containers; check existence",
        "When_NOT_To_Use": "Sorted range (use binary_search, lower_bound); conditional find (use find_if)",
        "Real_World_Freq": 9, "DSA_LeetCode_Freq": 9,
        "Example": "auto it = std::find(v.begin(), v.end(), 42); if(it != v.end()) {...}",
        "Notes": "Most common search; returns end() if not found",
        "Since_Version": "C++98", "Related": "find_if, binary_search"
    },
    {
        "Function": "std::find_if", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, UnaryPredicate p",
        "Arg_Explanation": "first, last: input range; p: predicate function",
        "Return_Type": "InputIt",
        "Description": "Finds first element satisfying predicate",
        "When_To_Use": "Conditional search; complex criteria; custom matching",
        "When_NOT_To_Use": "Simple value search (use find); sorted binary search",
        "Real_World_Freq": 9, "DSA_LeetCode_Freq": 10,
        "Example": "auto it = std::find_if(v.begin(), v.end(), [](int x){ return x > 10; });",
        "Notes": "Extremely common in DSA; essential for filtering",
        "Since_Version": "C++98", "Related": "find, find_if_not"
    },
    {
        "Function": "std::find_if_not", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, UnaryPredicate p",
        "Arg_Explanation": "first, last: input range; p: predicate function",
        "Return_Type": "InputIt",
        "Description": "Finds first element NOT satisfying predicate",
        "When_To_Use": "Find exception to rule; inverse condition",
        "When_NOT_To_Use": "Can negate predicate in find_if; simple case",
        "Real_World_Freq": 3, "DSA_LeetCode_Freq": 4,
        "Example": "auto it = std::find_if_not(v.begin(), v.end(), [](int x){ return x > 0; });",
        "Notes": "Equivalent to find_if with negated predicate",
        "Since_Version": "C++11", "Related": "find_if"
    },
    {
        "Function": "std::find_end", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n*m)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt1 first1, ForwardIt1 last1, ForwardIt2 first2, ForwardIt2 last2",
        "Arg_Explanation": "first1/last1: range to search in; first2/last2: subsequence to find",
        "Return_Type": "ForwardIt1",
        "Description": "Finds last occurrence of subsequence",
        "When_To_Use": "Find last match of pattern; reverse search",
        "When_NOT_To_Use": "First occurrence (use search); single element (use find)",
        "Real_World_Freq": 2, "DSA_LeetCode_Freq": 3,
        "Example": "auto it = std::find_end(text.begin(), text.end(), pattern.begin(), pattern.end());",
        "Notes": "Returns position of last occurrence or end()",
        "Since_Version": "C++98", "Related": "search, find_first_of"
    },
    {
        "Function": "std::find_first_of", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n*m)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt1 first1, InputIt1 last1, ForwardIt2 first2, ForwardIt2 last2",
        "Arg_Explanation": "first1/last1: range to search; first2/last2: values to search for",
        "Return_Type": "InputIt1",
        "Description": "Finds first occurrence of any element from second range",
        "When_To_Use": "Find any of multiple values; character set matching",
        "When_NOT_To_Use": "Single value (use find); set membership better with set",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 5,
        "Example": "auto it = std::find_first_of(str.begin(), str.end(), vowels.begin(), vowels.end());",
        "Notes": "Useful for string parsing; can be slow for large search sets",
        "Since_Version": "C++98", "Related": "find, search"
    },
    {
        "Function": "std::adjacent_find", "Header": "<algorithm>", "Category": "Non-modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, BinaryPredicate p",
        "Arg_Explanation": "first, last: input range; p: optional binary predicate",
        "Return_Type": "ForwardIt",
        "Description": "Finds first pair of adjacent equal elements",
        "When_To_Use": "Find consecutive duplicates; detect runs",
        "When_NOT_To_Use": "Non-adjacent duplicates; all duplicates",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 6,
        "Example": "auto it = std::adjacent_find(v.begin(), v.end());",
        "Notes": "Returns iterator to first of the pair",
        "Since_Version": "C++98", "Related": "unique, find"
    },
    # ========== SORTING ALGORITHMS ==========
    {
        "Function": "std::sort", "Header": "<algorithm>", "Category": "Sorting",
        "Time_Complexity": "O(n log n)", "Space_Complexity": "O(log n)",
        "Arguments": "RandomIt first, RandomIt last, Compare comp",
        "Arg_Explanation": "first, last: range to sort; comp: optional comparison function",
        "Return_Type": "void",
        "Description": "Sorts elements in ascending order (or by custom comparator)",
        "When_To_Use": "General sorting; most common sort; unstable sort OK",
        "When_NOT_To_Use": "Need stability (use stable_sort); partially sorted (use partial_sort); linked lists",
        "Real_World_Freq": 10, "DSA_LeetCode_Freq": 10,
        "Example": "std::sort(v.begin(), v.end()); // or with lambda: std::sort(v.begin(), v.end(), [](int a, int b){ return a > b; });",
        "Notes": "Usually IntroSort (QuickSort + HeapSort + InsertionSort); unstable; most used algorithm",
        "Since_Version": "C++98", "Related": "stable_sort, partial_sort, nth_element"
    },
    {
        "Function": "std::stable_sort", "Header": "<algorithm>", "Category": "Sorting",
        "Time_Complexity": "O(n log n)", "Space_Complexity": "O(n)",
        "Arguments": "RandomIt first, RandomIt last, Compare comp",
        "Arg_Explanation": "first, last: range to sort; comp: optional comparison function",
        "Return_Type": "void",
        "Description": "Sorts while preserving relative order of equal elements",
        "When_To_Use": "Need stability; multi-key sorting; preserve original order",
        "When_NOT_To_Use": "Stability not needed (use sort - faster); memory constrained",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 7,
        "Example": "std::stable_sort(people.begin(), people.end(), [](auto& a, auto& b){ return a.age < b.age; });",
        "Notes": "Usually MergeSort; requires extra memory; stable guarantee",
        "Since_Version": "C++98", "Related": "sort, partial_sort"
    },
    {
        "Function": "std::partial_sort", "Header": "<algorithm>", "Category": "Sorting",
        "Time_Complexity": "O(n log k)", "Space_Complexity": "O(1)",
        "Arguments": "RandomIt first, RandomIt middle, RandomIt last, Compare comp",
        "Arg_Explanation": "first: start; middle: end of sorted portion; last: end; comp: optional comparator",
        "Return_Type": "void",
        "Description": "Partially sorts so [first, middle) contains smallest elements sorted",
        "When_To_Use": "Top-k elements; only need first few sorted; k << n",
        "When_NOT_To_Use": "Need all elements sorted; k close to n (use sort)",
        "Real_World_Freq": 5, "DSA_LeetCode_Freq": 8,
        "Example": "std::partial_sort(v.begin(), v.begin() + 10, v.end()); // sort first 10",
        "Notes": "More efficient than full sort when k << n; uses heap",
        "Since_Version": "C++98", "Related": "nth_element, sort"
    },
    {
        "Function": "std::nth_element", "Header": "<algorithm>", "Category": "Sorting",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "RandomIt first, RandomIt nth, RandomIt last, Compare comp",
        "Arg_Explanation": "first, last: range; nth: position to partition around; comp: optional comparator",
        "Return_Type": "void",
        "Description": "Partitions so nth element is in sorted position, smaller before, larger after",
        "When_To_Use": "Find median; k-th smallest; quick select; partition",
        "When_NOT_To_Use": "Need sorted order; multiple k-th elements",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 9,
        "Example": "std::nth_element(v.begin(), v.begin() + v.size()/2, v.end()); // median at middle",
        "Notes": "Average O(n); QuickSelect; unstable; very useful for percentiles",
        "Since_Version": "C++98", "Related": "partial_sort, partition"
    },
    {
        "Function": "std::is_sorted", "Header": "<algorithm>", "Category": "Sorting",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, Compare comp",
        "Arg_Explanation": "first, last: range to check; comp: optional comparison function",
        "Return_Type": "bool",
        "Description": "Checks if range is sorted",
        "When_To_Use": "Validate sorted invariant; optimization check",
        "When_NOT_To_Use": "Need to know where unsorted; frequently checking",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 6,
        "Example": "if (std::is_sorted(v.begin(), v.end())) { /* use binary search */ }",
        "Notes": "Useful for assertions and optimizations",
        "Since_Version": "C++11", "Related": "is_sorted_until"
    },
    {
        "Function": "std::is_sorted_until", "Header": "<algorithm>", "Category": "Sorting",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, Compare comp",
        "Arg_Explanation": "first, last: range; comp: optional comparison function",
        "Return_Type": "ForwardIt",
        "Description": "Finds first position where range is no longer sorted",
        "When_To_Use": "Find where sorting breaks; partial sort validation",
        "When_NOT_To_Use": "Just need boolean (use is_sorted); don't need position",
        "Real_World_Freq": 2, "DSA_LeetCode_Freq": 3,
        "Example": "auto it = std::is_sorted_until(v.begin(), v.end());",
        "Notes": "Returns iterator to first out-of-order element or end()",
        "Since_Version": "C++11", "Related": "is_sorted"
    },
    
    # ========== BINARY SEARCH ALGORITHMS (on sorted ranges) ==========
    {
        "Function": "std::binary_search", "Header": "<algorithm>", "Category": "Binary Search",
        "Time_Complexity": "O(log n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, const T& value, Compare comp",
        "Arg_Explanation": "first, last: SORTED range; value: value to find; comp: optional comparator",
        "Return_Type": "bool",
        "Description": "Checks if value exists in sorted range",
        "When_To_Use": "Existence check in sorted data; fast membership test",
        "When_NOT_To_Use": "Unsorted data; need iterator to element (use lower_bound); need count",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 9,
        "Example": "bool found = std::binary_search(v.begin(), v.end(), 42);",
        "Notes": "REQUIRES sorted range; only returns bool; use lower_bound for iterator",
        "Since_Version": "C++98", "Related": "lower_bound, upper_bound, equal_range"
    },
    {
        "Function": "std::lower_bound", "Header": "<algorithm>", "Category": "Binary Search",
        "Time_Complexity": "O(log n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, const T& value, Compare comp",
        "Arg_Explanation": "first, last: SORTED range; value: value to find; comp: optional comparator",
        "Return_Type": "ForwardIt",
        "Description": "Finds first element not less than value (>=)",
        "When_To_Use": "Find insertion point; first >= value; range queries",
        "When_NOT_To_Use": "Unsorted data; just existence check (use binary_search)",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 10,
        "Example": "auto it = std::lower_bound(v.begin(), v.end(), 42); // first element >= 42",
        "Notes": "ESSENTIAL for DSA; returns end() if all < value; insertion point",
        "Since_Version": "C++98", "Related": "upper_bound, equal_range, binary_search"
    },
    {
        "Function": "std::upper_bound", "Header": "<algorithm>", "Category": "Binary Search",
        "Time_Complexity": "O(log n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, const T& value, Compare comp",
        "Arg_Explanation": "first, last: SORTED range; value: value to find; comp: optional comparator",
        "Return_Type": "ForwardIt",
        "Description": "Finds first element greater than value (>)",
        "When_To_Use": "Find insertion point after value; first > value; range queries",
        "When_NOT_To_Use": "Unsorted data; need >= (use lower_bound)",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 10,
        "Example": "auto it = std::upper_bound(v.begin(), v.end(), 42); // first element > 42",
        "Notes": "Combined with lower_bound gives range of equal elements",
        "Since_Version": "C++98", "Related": "lower_bound, equal_range"
    },
    {
        "Function": "std::equal_range", "Header": "<algorithm>", "Category": "Binary Search",
        "Time_Complexity": "O(log n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, const T& value, Compare comp",
        "Arg_Explanation": "first, last: SORTED range; value: value to find; comp: optional comparator",
        "Return_Type": "pair<ForwardIt, ForwardIt>",
        "Description": "Returns range [lower_bound, upper_bound) of equal elements",
        "When_To_Use": "Find all occurrences of value; range of equals",
        "When_NOT_To_Use": "Just existence; single occurrence expected",
        "Real_World_Freq": 5, "DSA_LeetCode_Freq": 7,
        "Example": "auto [first, last] = std::equal_range(v.begin(), v.end(), 42);",
        "Notes": "Equivalent to {lower_bound, upper_bound}; useful for counting",
        "Since_Version": "C++98", "Related": "lower_bound, upper_bound"
    },
    # ========== MODIFYING SEQUENCE OPERATIONS ==========
    {
        "Function": "std::copy", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, OutputIt d_first",
        "Arg_Explanation": "first, last: source range; d_first: destination start",
        "Return_Type": "OutputIt",
        "Description": "Copies elements from source to destination",
        "When_To_Use": "Copy range to another container; clone data",
        "When_NOT_To_Use": "Overlapping ranges (use copy_backward); transformation needed (use transform)",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 6,
        "Example": "std::copy(src.begin(), src.end(), dest.begin());",
        "Notes": "Destination must have space; returns iterator past last copied",
        "Since_Version": "C++98", "Related": "copy_if, copy_n, copy_backward"
    },
    {
        "Function": "std::copy_if", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, OutputIt d_first, UnaryPredicate p",
        "Arg_Explanation": "first, last: source; d_first: destination; p: predicate for filtering",
        "Return_Type": "OutputIt",
        "Description": "Copies elements satisfying predicate",
        "When_To_Use": "Filtered copy; conditional data transfer",
        "When_NOT_To_Use": "All elements (use copy); in-place filtering (use remove_if)",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 8,
        "Example": "std::copy_if(src.begin(), src.end(), std::back_inserter(dest), [](int x){ return x > 0; });",
        "Notes": "Very useful with back_inserter; common in filtering",
        "Since_Version": "C++11", "Related": "copy, remove_copy_if"
    },
    {
        "Function": "std::copy_n", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, Size count, OutputIt d_first",
        "Arg_Explanation": "first: source start; count: number of elements; d_first: destination",
        "Return_Type": "OutputIt",
        "Description": "Copies exactly n elements",
        "When_To_Use": "Copy specific count; array-like copying",
        "When_NOT_To_Use": "Copy entire range (use copy); conditional copy",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 4,
        "Example": "std::copy_n(src.begin(), 10, dest.begin());",
        "Notes": "Doesn't check bounds; efficient for fixed counts",
        "Since_Version": "C++11", "Related": "copy"
    },
    {
        "Function": "std::fill", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, const T& value",
        "Arg_Explanation": "first, last: range to fill; value: value to assign",
        "Return_Type": "void",
        "Description": "Assigns value to all elements in range",
        "When_To_Use": "Initialize range; reset values; set all to same value",
        "When_NOT_To_Use": "Need different values (use generate); construction (use vector constructor)",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 7,
        "Example": "std::fill(v.begin(), v.end(), 0); // set all to 0",
        "Notes": "Common for initialization; simple and clear",
        "Since_Version": "C++98", "Related": "fill_n, generate"
    },
    {
        "Function": "std::fill_n", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "OutputIt first, Size count, const T& value",
        "Arg_Explanation": "first: start position; count: number of elements; value: value to assign",
        "Return_Type": "OutputIt",
        "Description": "Assigns value to first n elements",
        "When_To_Use": "Fill specific count; array initialization",
        "When_NOT_To_Use": "Fill entire range (use fill); different values per element",
        "Real_World_Freq": 5, "DSA_LeetCode_Freq": 5,
        "Example": "std::fill_n(v.begin(), 10, -1); // first 10 elements to -1",
        "Notes": "Returns iterator past last filled element",
        "Since_Version": "C++98", "Related": "fill"
    },
    {
        "Function": "std::transform", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, OutputIt d_first, UnaryOp op",
        "Arg_Explanation": "first, last: source; d_first: destination; op: transformation function",
        "Return_Type": "OutputIt",
        "Description": "Applies function to range and stores result",
        "When_To_Use": "Map operation; element-wise transformation; functional programming",
        "When_NOT_To_Use": "No transformation (use copy); in-place with same type (can use transform with same it)",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 8,
        "Example": "std::transform(v.begin(), v.end(), v.begin(), [](int x){ return x * 2; });",
        "Notes": "Can transform in-place; binary version for two inputs",
        "Since_Version": "C++98", "Related": "for_each, copy"
    },
    {
        "Function": "std::generate", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, Generator g",
        "Arg_Explanation": "first, last: range; g: generator function (no args, returns value)",
        "Return_Type": "void",
        "Description": "Assigns values generated by function",
        "When_To_Use": "Generate sequence; random values; computed values",
        "When_NOT_To_Use": "Same value for all (use fill); transformation of existing (use transform)",
        "Real_World_Freq": 5, "DSA_LeetCode_Freq": 4,
        "Example": "std::generate(v.begin(), v.end(), std::rand);",
        "Notes": "Generator called once per element; useful for random/sequential data",
        "Since_Version": "C++98", "Related": "generate_n, fill"
    },
    {
        "Function": "std::remove", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, const T& value",
        "Arg_Explanation": "first, last: range; value: value to remove",
        "Return_Type": "ForwardIt",
        "Description": "Removes elements equal to value (moves to end)",
        "When_To_Use": "Remove specific value; erase-remove idiom",
        "When_NOT_To_Use": "Conditional removal (use remove_if); actually erase (combine with erase)",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 8,
        "Example": "v.erase(std::remove(v.begin(), v.end(), 42), v.end()); // erase-remove idiom",
        "Notes": "Doesn't actually erase; returns new logical end; use erase-remove idiom",
        "Since_Version": "C++98", "Related": "remove_if, erase"
    },
    {
        "Function": "std::remove_if", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, UnaryPredicate p",
        "Arg_Explanation": "first, last: range; p: predicate for removal condition",
        "Return_Type": "ForwardIt",
        "Description": "Removes elements satisfying predicate",
        "When_To_Use": "Conditional removal; filter out elements; erase-remove idiom",
        "When_NOT_To_Use": "Simple value (use remove); keep matching (invert predicate or use copy_if)",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 9,
        "Example": "v.erase(std::remove_if(v.begin(), v.end(), [](int x){ return x < 0; }), v.end());",
        "Notes": "VERY common in DSA; erase-remove idiom essential pattern",
        "Since_Version": "C++98", "Related": "remove, erase, copy_if"
    },
    {
        "Function": "std::unique", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, BinaryPredicate p",
        "Arg_Explanation": "first, last: range (should be sorted); p: optional equality predicate",
        "Return_Type": "ForwardIt",
        "Description": "Removes consecutive duplicate elements",
        "When_To_Use": "Remove adjacent duplicates; after sorting for all duplicates",
        "When_NOT_To_Use": "Need all unique (sort first); non-adjacent duplicates",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 8,
        "Example": "std::sort(v.begin(), v.end()); v.erase(std::unique(v.begin(), v.end()), v.end());",
        "Notes": "Only removes CONSECUTIVE duplicates; sort first for all duplicates",
        "Since_Version": "C++98", "Related": "remove, sort"
    },
    {
        "Function": "std::reverse", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "BidirectionalIt first, BidirectionalIt last",
        "Arg_Explanation": "first, last: range to reverse",
        "Return_Type": "void",
        "Description": "Reverses order of elements in range",
        "When_To_Use": "Reverse array/vector; palindrome check; reverse iteration order",
        "When_NOT_To_Use": "Temporary reverse iteration (use reverse_iterator); copy reversed (use reverse_copy)",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 8,
        "Example": "std::reverse(v.begin(), v.end());",
        "Notes": "In-place; O(1) space; bidirectional iterators required",
        "Since_Version": "C++98", "Related": "reverse_copy, rotate"
    },
    {
        "Function": "std::rotate", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt n_first, ForwardIt last",
        "Arg_Explanation": "first: range start; n_first: new first element; last: range end",
        "Return_Type": "ForwardIt",
        "Description": "Rotates elements so n_first becomes first",
        "When_To_Use": "Circular shift; move elements; rearrange",
        "When_NOT_To_Use": "Simple swap; sort (use sort)",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 6,
        "Example": "std::rotate(v.begin(), v.begin() + 3, v.end()); // [0,1,2,3,4] -> [3,4,0,1,2]",
        "Notes": "Useful for array rotation problems; O(1) space",
        "Since_Version": "C++98", "Related": "reverse, swap_ranges"
    },
    {
        "Function": "std::shuffle", "Header": "<algorithm>", "Category": "Modifying",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "RandomIt first, RandomIt last, URBG&& g",
        "Arg_Explanation": "first, last: range; g: random number generator",
        "Return_Type": "void",
        "Description": "Randomly reorders elements",
        "When_To_Use": "Randomize order; shuffle deck; random sampling",
        "When_NOT_To_Use": "Need specific permutation; deterministic reorder",
        "Real_World_Freq": 5, "DSA_LeetCode_Freq": 4,
        "Example": "std::shuffle(v.begin(), v.end(), std::mt19937{std::random_device{}()});",
        "Notes": "Replaces deprecated random_shuffle; requires C++11 random generators",
        "Since_Version": "C++11", "Related": "random_shuffle (deprecated)"
    },
    # ========== HEAP OPERATIONS ==========
    {
        "Function": "std::make_heap", "Header": "<algorithm>", "Category": "Heap",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "RandomIt first, RandomIt last, Compare comp",
        "Arg_Explanation": "first, last: range to heapify; comp: optional comparison function",
        "Return_Type": "void",
        "Description": "Converts range into max heap",
        "When_To_Use": "Create heap from array; priority queue implementation",
        "When_NOT_To_Use": "Need container-based pq (use priority_queue); frequently changing",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 6,
        "Example": "std::make_heap(v.begin(), v.end());",
        "Notes": "Linear time construction; max heap by default",
        "Since_Version": "C++98", "Related": "push_heap, pop_heap, priority_queue"
    },
    {
        "Function": "std::push_heap", "Header": "<algorithm>", "Category": "Heap",
        "Time_Complexity": "O(log n)", "Space_Complexity": "O(1)",
        "Arguments": "RandomIt first, RandomIt last, Compare comp",
        "Arg_Explanation": "first, last: heap range (last-1 is new element); comp: optional comparator",
        "Return_Type": "void",
        "Description": "Inserts element at end into heap",
        "When_To_Use": "Add to existing heap; manual heap operations",
        "When_NOT_To_Use": "Use priority_queue for easier interface",
        "Real_World_Freq": 3, "DSA_LeetCode_Freq": 5,
        "Example": "v.push_back(42); std::push_heap(v.begin(), v.end());",
        "Notes": "Element must be at end before calling; maintains heap property",
        "Since_Version": "C++98", "Related": "make_heap, pop_heap"
    },
    {
        "Function": "std::pop_heap", "Header": "<algorithm>", "Category": "Heap",
        "Time_Complexity": "O(log n)", "Space_Complexity": "O(1)",
        "Arguments": "RandomIt first, RandomIt last, Compare comp",
        "Arg_Explanation": "first, last: heap range; comp: optional comparator",
        "Return_Type": "void",
        "Description": "Moves largest element to end, maintains heap for rest",
        "When_To_Use": "Extract max from heap; manual heap operations",
        "When_NOT_To_Use": "Use priority_queue for easier interface",
        "Real_World_Freq": 3, "DSA_LeetCode_Freq": 5,
        "Example": "std::pop_heap(v.begin(), v.end()); int max = v.back(); v.pop_back();",
        "Notes": "Max element moved to end; must manually remove it",
        "Since_Version": "C++98", "Related": "push_heap, make_heap"
    },
    {
        "Function": "std::sort_heap", "Header": "<algorithm>", "Category": "Heap",
        "Time_Complexity": "O(n log n)", "Space_Complexity": "O(1)",
        "Arguments": "RandomIt first, RandomIt last, Compare comp",
        "Arg_Explanation": "first, last: heap range; comp: optional comparator",
        "Return_Type": "void",
        "Description": "Converts heap to sorted range",
        "When_To_Use": "Heap sort; convert heap to sorted array",
        "When_NOT_To_Use": "Don't need sorted (keep as heap); use regular sort",
        "Real_World_Freq": 2, "DSA_LeetCode_Freq": 3,
        "Example": "std::make_heap(v.begin(), v.end()); std::sort_heap(v.begin(), v.end());",
        "Notes": "Range is no longer a heap after; in-place heap sort",
        "Since_Version": "C++98", "Related": "make_heap, sort"
    },
    
    # ========== SET OPERATIONS (on sorted ranges) ==========
    {
        "Function": "std::set_union", "Header": "<algorithm>", "Category": "Set Operations",
        "Time_Complexity": "O(n + m)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt1 first1, InputIt1 last1, InputIt2 first2, InputIt2 last2, OutputIt d_first",
        "Arg_Explanation": "first1/last1, first2/last2: two SORTED ranges; d_first: output destination",
        "Return_Type": "OutputIt",
        "Description": "Computes union of two sorted ranges",
        "When_To_Use": "Merge sorted sets; combine unique elements",
        "When_NOT_To_Use": "Unsorted data; unordered_set better for sets",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 6,
        "Example": "std::set_union(v1.begin(), v1.end(), v2.begin(), v2.end(), std::back_inserter(result));",
        "Notes": "BOTH ranges must be sorted; includes duplicates based on max count",
        "Since_Version": "C++98", "Related": "set_intersection, set_difference, merge"
    },
    {
        "Function": "std::set_intersection", "Header": "<algorithm>", "Category": "Set Operations",
        "Time_Complexity": "O(n + m)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt1 first1, InputIt1 last1, InputIt2 first2, InputIt2 last2, OutputIt d_first",
        "Arg_Explanation": "first1/last1, first2/last2: two SORTED ranges; d_first: output",
        "Return_Type": "OutputIt",
        "Description": "Computes intersection of two sorted ranges",
        "When_To_Use": "Find common elements; set intersection",
        "When_NOT_To_Use": "Unsorted data; hash set intersection easier",
        "Real_World_Freq": 5, "DSA_LeetCode_Freq": 7,
        "Example": "std::set_intersection(v1.begin(), v1.end(), v2.begin(), v2.end(), std::back_inserter(result));",
        "Notes": "BOTH ranges must be sorted; common LeetCode pattern",
        "Since_Version": "C++98", "Related": "set_union, set_difference"
    },
    {
        "Function": "std::set_difference", "Header": "<algorithm>", "Category": "Set Operations",
        "Time_Complexity": "O(n + m)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt1 first1, InputIt1 last1, InputIt2 first2, InputIt2 last2, OutputIt d_first",
        "Arg_Explanation": "first1/last1, first2/last2: two SORTED ranges; d_first: output",
        "Return_Type": "OutputIt",
        "Description": "Computes set difference (elements in first but not second)",
        "When_To_Use": "Find elements in A but not B; set subtraction",
        "When_NOT_To_Use": "Unsorted data; order doesn't matter",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 6,
        "Example": "std::set_difference(v1.begin(), v1.end(), v2.begin(), v2.end(), std::back_inserter(result));",
        "Notes": "BOTH ranges must be sorted; not symmetric",
        "Since_Version": "C++98", "Related": "set_symmetric_difference"
    },
    {
        "Function": "std::merge", "Header": "<algorithm>", "Category": "Set Operations",
        "Time_Complexity": "O(n + m)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt1 first1, InputIt1 last1, InputIt2 first2, InputIt2 last2, OutputIt d_first",
        "Arg_Explanation": "first1/last1, first2/last2: two SORTED ranges; d_first: output",
        "Return_Type": "OutputIt",
        "Description": "Merges two sorted ranges into one sorted range",
        "When_To_Use": "Merge sort; combine sorted sequences; stable merge",
        "When_NOT_To_Use": "Unsorted data (sort first); in-place merge",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 8,
        "Example": "std::merge(v1.begin(), v1.end(), v2.begin(), v2.end(), std::back_inserter(result));",
        "Notes": "Essential for merge sort; stable; preserves duplicates",
        "Since_Version": "C++98", "Related": "inplace_merge, set_union"
    },
    {
        "Function": "std::includes", "Header": "<algorithm>", "Category": "Set Operations",
        "Time_Complexity": "O(n + m)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt1 first1, InputIt1 last1, InputIt2 first2, InputIt2 last2",
        "Arg_Explanation": "first1/last1: set to check; first2/last2: subset to find",
        "Return_Type": "bool",
        "Description": "Checks if one sorted range is subset of another",
        "When_To_Use": "Subset verification; containment check",
        "When_NOT_To_Use": "Unsorted data; equality check (use equal)",
        "Real_World_Freq": 3, "DSA_LeetCode_Freq": 5,
        "Example": "bool is_subset = std::includes(set.begin(), set.end(), subset.begin(), subset.end());",
        "Notes": "BOTH ranges must be sorted",
        "Since_Version": "C++98", "Related": "set_intersection"
    },
    
    # ========== MIN/MAX OPERATIONS ==========
    {
        "Function": "std::max", "Header": "<algorithm>", "Category": "Min/Max",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "const T& a, const T& b, Compare comp",
        "Arg_Explanation": "a, b: values to compare; comp: optional comparator",
        "Return_Type": "const T&",
        "Description": "Returns larger of two values",
        "When_To_Use": "Compare two values; simple maximum",
        "When_NOT_To_Use": "More than 2 values (use max with initializer_list); range (use max_element)",
        "Real_World_Freq": 10, "DSA_LeetCode_Freq": 10,
        "Example": "int m = std::max(a, b); int m2 = std::max({1, 5, 3, 2}); // C++11",
        "Notes": "Extremely common; has initializer_list version in C++11",
        "Since_Version": "C++98", "Related": "min, max_element, clamp"
    },
    {
        "Function": "std::min", "Header": "<algorithm>", "Category": "Min/Max",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "const T& a, const T& b, Compare comp",
        "Arg_Explanation": "a, b: values to compare; comp: optional comparator",
        "Return_Type": "const T&",
        "Description": "Returns smaller of two values",
        "When_To_Use": "Compare two values; simple minimum",
        "When_NOT_To_Use": "More than 2 values (use initializer_list version); range (use min_element)",
        "Real_World_Freq": 10, "DSA_LeetCode_Freq": 10,
        "Example": "int m = std::min(a, b); int m2 = std::min({1, 5, 3, 2}); // C++11",
        "Notes": "Extremely common; has initializer_list version in C++11",
        "Since_Version": "C++98", "Related": "max, min_element, clamp"
    },
    {
        "Function": "std::minmax", "Header": "<algorithm>", "Category": "Min/Max",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "const T& a, const T& b, Compare comp",
        "Arg_Explanation": "a, b: values to compare; comp: optional comparator",
        "Return_Type": "pair<const T&, const T&>",
        "Description": "Returns pair of min and max",
        "When_To_Use": "Need both min and max; single comparison",
        "When_NOT_To_Use": "Only need one; range (use minmax_element)",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 5,
        "Example": "auto [minimum, maximum] = std::minmax(a, b);",
        "Notes": "More efficient than calling min and max separately",
        "Since_Version": "C++11", "Related": "min, max, minmax_element"
    },
    {
        "Function": "std::max_element", "Header": "<algorithm>", "Category": "Min/Max",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, Compare comp",
        "Arg_Explanation": "first, last: range to search; comp: optional comparator",
        "Return_Type": "ForwardIt",
        "Description": "Finds iterator to largest element in range",
        "When_To_Use": "Find maximum in container; get max iterator",
        "When_NOT_To_Use": "Just need value (dereference result); two values (use max)",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 9,
        "Example": "auto it = std::max_element(v.begin(), v.end()); int max_val = *it;",
        "Notes": "Returns iterator, not value; very common in DSA",
        "Since_Version": "C++98", "Related": "min_element, minmax_element"
    },
    {
        "Function": "std::min_element", "Header": "<algorithm>", "Category": "Min/Max",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, Compare comp",
        "Arg_Explanation": "first, last: range to search; comp: optional comparator",
        "Return_Type": "ForwardIt",
        "Description": "Finds iterator to smallest element in range",
        "When_To_Use": "Find minimum in container; get min iterator",
        "When_NOT_To_Use": "Just need value (dereference result); two values (use min)",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 9,
        "Example": "auto it = std::min_element(v.begin(), v.end()); int min_val = *it;",
        "Notes": "Returns iterator, not value; very common in DSA",
        "Since_Version": "C++98", "Related": "max_element, minmax_element"
    },
    {
        "Function": "std::clamp", "Header": "<algorithm>", "Category": "Min/Max",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "const T& v, const T& lo, const T& hi, Compare comp",
        "Arg_Explanation": "v: value to clamp; lo: lower bound; hi: upper bound; comp: optional comparator",
        "Return_Type": "const T&",
        "Description": "Clamps value between lower and upper bounds",
        "When_To_Use": "Bound value to range; sanitize input; constrain values",
        "When_NOT_To_Use": "Simple min/max; no bounds needed",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 5,
        "Example": "int clamped = std::clamp(value, 0, 100); // ensures 0 <= clamped <= 100",
        "Notes": "Very useful for bounds checking; added in C++17",
        "Since_Version": "C++17", "Related": "min, max"
    },
    # ========== PERMUTATION & PARTITIONING ==========
    {
        "Function": "std::next_permutation", "Header": "<algorithm>", "Category": "Permutation",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "BidirectionalIt first, BidirectionalIt last",
        "Arg_Explanation": "first, last: range to permute",
        "Return_Type": "bool",
        "Description": "Transforms range into next lexicographically greater permutation",
        "When_To_Use": "Generate all permutations; combinatorial problems",
        "When_NOT_To_Use": "Random permutation (use shuffle)",
        "Real_World_Freq": 3, "DSA_LeetCode_Freq": 9,
        "Example": "do { process(v); } while (std::next_permutation(v.begin(), v.end()));",
        "Notes": "Essential for permutation problems; start with sorted range",
        "Since_Version": "C++98", "Related": "prev_permutation"
    },
    {
        "Function": "std::partition", "Header": "<algorithm>", "Category": "Partitioning",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, UnaryPredicate p",
        "Arg_Explanation": "first, last: range; p: predicate",
        "Return_Type": "ForwardIt",
        "Description": "Reorders so predicate-true elements come first",
        "When_To_Use": "Separate by condition; quick select",
        "When_NOT_To_Use": "Need stability (use stable_partition)",
        "Real_World_Freq": 5, "DSA_LeetCode_Freq": 7,
        "Example": "auto mid = std::partition(v.begin(), v.end(), [](int x){ return x % 2 == 0; });",
        "Notes": "Unstable; used in quicksort/quickselect",
        "Since_Version": "C++98", "Related": "stable_partition, nth_element"
    },
    {
        "Function": "std::accumulate", "Header": "<numeric>", "Category": "Numeric",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "InputIt first, InputIt last, T init",
        "Arg_Explanation": "first, last: range; init: initial value",
        "Return_Type": "T",
        "Description": "Computes sum (or fold) of range",
        "When_To_Use": "Sum/product elements; fold operations",
        "When_NOT_To_Use": "Just counting (use distance)",
        "Real_World_Freq": 9, "DSA_LeetCode_Freq": 9,
        "Example": "int sum = std::accumulate(v.begin(), v.end(), 0);",
        "Notes": "Most common aggregation; can use custom operation",
        "Since_Version": "C++98", "Related": "reduce, transform_reduce"
    },
    {
        "Function": "std::iota", "Header": "<numeric>", "Category": "Numeric",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "ForwardIt first, ForwardIt last, T value",
        "Arg_Explanation": "first, last: range; value: starting value",
        "Return_Type": "void",
        "Description": "Fills range with sequentially increasing values",
        "When_To_Use": "Generate sequence 0,1,2...; index arrays",
        "When_NOT_To_Use": "Complex sequence (use generate)",
        "Real_World_Freq": 5, "DSA_LeetCode_Freq": 7,
        "Example": "std::iota(v.begin(), v.end(), 0); // 0,1,2,3...",
        "Notes": "Simple and elegant for sequences",
        "Since_Version": "C++11", "Related": "generate"
    },
]


# ==============================================================================
# VECTOR METHODS - std::vector<T>
# ==============================================================================
vector_methods = [
    # === ELEMENT ACCESS ===
    {
        "Container": "vector", "Method": "operator[]", "Category": "Element Access",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "size_t pos",
        "Arg_Explanation": "pos: index of element",
        "Return_Type": "T&",
        "Description": "Access element at index (no bounds checking)",
        "When_To_Use": "Fast access; index guaranteed valid",
        "When_NOT_To_Use": "Need bounds checking (use at()); unknown index validity",
        "Real_World_Freq": 10, "DSA_LeetCode_Freq": 10,
        "Example": "int val = v[5];",
        "Notes": "Most common access; undefined behavior if out of bounds",
        "Since_Version": "C++98", "Related": "at, front, back"
    },
    {
        "Container": "vector", "Method": "at", "Category": "Element Access",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "size_t pos",
        "Arg_Explanation": "pos: index of element",
        "Return_Type": "T&",
        "Description": "Access element with bounds checking (throws if invalid)",
        "When_To_Use": "Need safety; untrusted indices",
        "When_NOT_To_Use": "Performance critical; index guaranteed valid",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 4,
        "Example": "try { int val = v.at(5); } catch(std::out_of_range& e) {}",
        "Notes": "Throws out_of_range exception; safer than []",
        "Since_Version": "C++98", "Related": "operator[]"
    },
    {
        "Container": "vector", "Method": "front", "Category": "Element Access",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "none",
        "Arg_Explanation": "",
        "Return_Type": "T&",
        "Description": "Access first element",
        "When_To_Use": "Get/modify first element",
        "When_NOT_To_Use": "Empty vector (undefined behavior)",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 8,
        "Example": "int first = v.front();",
        "Notes": "Undefined if empty; check !v.empty() first",
        "Since_Version": "C++98", "Related": "back, operator[]"
    },
    {
        "Container": "vector", "Method": "back", "Category": "Element Access",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "none",
        "Arg_Explanation": "",
        "Return_Type": "T&",
        "Description": "Access last element",
        "When_To_Use": "Get/modify last element",
        "When_NOT_To_Use": "Empty vector (undefined behavior)",
        "Real_World_Freq": 9, "DSA_LeetCode_Freq": 9,
        "Example": "int last = v.back();",
        "Notes": "Very common; undefined if empty",
        "Since_Version": "C++98", "Related": "front, pop_back"
    },
    {
        "Container": "vector", "Method": "data", "Category": "Element Access",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "none",
        "Arg_Explanation": "",
        "Return_Type": "T*",
        "Description": "Returns pointer to underlying array",
        "When_To_Use": "C API interop; direct memory access",
        "When_NOT_To_Use": "Normal iteration (use iterators)",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 3,
        "Example": "int* ptr = v.data(); some_c_function(ptr, v.size());",
        "Notes": "Useful for C interop; contiguous memory guaranteed",
        "Since_Version": "C++11", "Related": "begin, operator[]"
    },
    
    # === CAPACITY ===
    {
        "Container": "vector", "Method": "size", "Category": "Capacity",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "none",
        "Arg_Explanation": "",
        "Return_Type": "size_t",
        "Description": "Returns number of elements",
        "When_To_Use": "Check size; loop bounds; allocation",
        "When_NOT_To_Use": "Check if empty (use empty())",
        "Real_World_Freq": 10, "DSA_LeetCode_Freq": 10,
        "Example": "for (size_t i = 0; i < v.size(); ++i) {}",
        "Notes": "Extremely common; constant time",
        "Since_Version": "C++98", "Related": "empty, capacity"
    },
    {
        "Container": "vector", "Method": "empty", "Category": "Capacity",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "none",
        "Arg_Explanation": "",
        "Return_Type": "bool",
        "Description": "Checks if container is empty",
        "When_To_Use": "Check emptiness; validation",
        "When_NOT_To_Use": "Need size (use size())",
        "Real_World_Freq": 9, "DSA_LeetCode_Freq": 9,
        "Example": "if (!v.empty()) { process(v.front()); }",
        "Notes": "Preferred over size() == 0; more expressive",
        "Since_Version": "C++98", "Related": "size"
    },
    {
        "Container": "vector", "Method": "reserve", "Category": "Capacity",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(n)",
        "Arguments": "size_t new_cap",
        "Arg_Explanation": "new_cap: new capacity",
        "Return_Type": "void",
        "Description": "Reserves storage to avoid reallocations",
        "When_To_Use": "Know final size; optimize allocations; avoid invalidation",
        "When_NOT_To_Use": "Size unknown; memory constrained",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 5,
        "Example": "v.reserve(1000); // pre-allocate for 1000 elements",
        "Notes": "Important optimization; prevents iterator invalidation",
        "Since_Version": "C++98", "Related": "capacity, resize, shrink_to_fit"
    },
    {
        "Container": "vector", "Method": "capacity", "Category": "Capacity",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "none",
        "Arg_Explanation": "",
        "Return_Type": "size_t",
        "Description": "Returns allocated storage capacity",
        "When_To_Use": "Check allocation; understand memory",
        "When_NOT_To_Use": "Normal usage (internal detail)",
        "Real_World_Freq": 4, "DSA_LeetCode_Freq": 2,
        "Example": "size_t cap = v.capacity();",
        "Notes": "Always >= size(); implementation defined growth",
        "Since_Version": "C++98", "Related": "size, reserve"
    },
    {
        "Container": "vector", "Method": "resize", "Category": "Capacity",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(n)",
        "Arguments": "size_t count, const T& value",
        "Arg_Explanation": "count: new size; value: optional value for new elements",
        "Return_Type": "void",
        "Description": "Changes size, adding/removing elements as needed",
        "When_To_Use": "Change size; initialize to specific size",
        "When_NOT_To_Use": "Just reserve capacity (use reserve)",
        "Real_World_Freq": 7, "DSA_LeetCode_Freq": 7,
        "Example": "v.resize(100); v.resize(50, -1);",
        "Notes": "Can grow or shrink; new elements default-constructed",
        "Since_Version": "C++98", "Related": "reserve, size"
    },
    
    # === MODIFIERS ===
    {
        "Container": "vector", "Method": "push_back", "Category": "Modifiers",
        "Time_Complexity": "Amortized O(1)", "Space_Complexity": "O(1)",
        "Arguments": "const T& value or T&& value",
        "Arg_Explanation": "value: element to add",
        "Return_Type": "void",
        "Description": "Adds element to end",
        "When_To_Use": "Build vector; add elements",
        "When_NOT_To_Use": "Frequent front insertion (use deque)",
        "Real_World_Freq": 10, "DSA_LeetCode_Freq": 10,
        "Example": "v.push_back(42);",
        "Notes": "Most common operation; amortized O(1)",
        "Since_Version": "C++98", "Related": "emplace_back, pop_back"
    },
    {
        "Container": "vector", "Method": "emplace_back", "Category": "Modifiers",
        "Time_Complexity": "Amortized O(1)", "Space_Complexity": "O(1)",
        "Arguments": "Args&&... args",
        "Arg_Explanation": "args: arguments to forward to constructor",
        "Return_Type": "T& (C++17+)",
        "Description": "Constructs element in-place at end",
        "When_To_Use": "Avoid copy; construct directly; efficiency",
        "When_NOT_To_Use": "Simple types where copy is cheap",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 6,
        "Example": "v.emplace_back(arg1, arg2); // constructs T(arg1, arg2)",
        "Notes": "More efficient than push_back for complex types",
        "Since_Version": "C++11", "Related": "push_back, emplace"
    },
    {
        "Container": "vector", "Method": "pop_back", "Category": "Modifiers",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "none",
        "Arg_Explanation": "",
        "Return_Type": "void",
        "Description": "Removes last element",
        "When_To_Use": "Remove from end; stack operations",
        "When_NOT_To_Use": "Empty vector (undefined); need value (get with back() first)",
        "Real_World_Freq": 9, "DSA_LeetCode_Freq": 9,
        "Example": "if (!v.empty()) v.pop_back();",
        "Notes": "Doesn't return value; check empty first",
        "Since_Version": "C++98", "Related": "push_back, back"
    },
    {
        "Container": "vector", "Method": "insert", "Category": "Modifiers",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "const_iterator pos, const T& value",
        "Arg_Explanation": "pos: position to insert; value: element to insert",
        "Return_Type": "iterator",
        "Description": "Inserts element(s) before position",
        "When_To_Use": "Insert in middle; specific position",
        "When_NOT_To_Use": "Frequent inserts (use list); at end (use push_back)",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 6,
        "Example": "v.insert(v.begin() + 5, 42); v.insert(v.end(), 3, 99);",
        "Notes": "Expensive O(n); shifts elements; invalidates iterators",
        "Since_Version": "C++98", "Related": "erase, emplace"
    },
    {
        "Container": "vector", "Method": "erase", "Category": "Modifiers",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "const_iterator pos or const_iterator first, const_iterator last",
        "Arg_Explanation": "pos: element to remove; or first/last: range to remove",
        "Return_Type": "iterator",
        "Description": "Removes element(s) from container",
        "When_To_Use": "Remove by position; erase-remove idiom",
        "When_NOT_To_Use": "Remove from end (use pop_back); by value (use erase-remove)",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 9,
        "Example": "v.erase(v.begin() + 5); v.erase(std::remove(v.begin(), v.end(), 42), v.end());",
        "Notes": "Erase-remove idiom essential; shifts elements",
        "Since_Version": "C++98", "Related": "remove, remove_if"
    },
    {
        "Container": "vector", "Method": "clear", "Category": "Modifiers",
        "Time_Complexity": "O(n)", "Space_Complexity": "O(1)",
        "Arguments": "none",
        "Arg_Explanation": "",
        "Return_Type": "void",
        "Description": "Removes all elements",
        "When_To_Use": "Empty container; reset state",
        "When_NOT_To_Use": "Need to free memory (doesn't affect capacity)",
        "Real_World_Freq": 8, "DSA_LeetCode_Freq": 7,
        "Example": "v.clear();",
        "Notes": "Size becomes 0; capacity unchanged",
        "Since_Version": "C++98", "Related": "erase, shrink_to_fit"
    },
    {
        "Container": "vector", "Method": "swap", "Category": "Modifiers",
        "Time_Complexity": "O(1)", "Space_Complexity": "O(1)",
        "Arguments": "vector& other",
        "Arg_Explanation": "other: vector to swap with",
        "Return_Type": "void",
        "Description": "Swaps contents with another vector",
        "When_To_Use": "Exchange vectors; move semantics",
        "When_NOT_To_Use": "Copy needed (use assignment)",
        "Real_World_Freq": 6, "DSA_LeetCode_Freq": 5,
        "Example": "v1.swap(v2); std::swap(v1, v2);",
        "Notes": "Constant time; swaps pointers; iterators remain valid to swapped container",
        "Since_Version": "C++98", "Related": "std::swap"
    },
]


# ==============================================================================
# MAP METHODS - std::map<K,V> and std::unordered_map<K,V>
# ==============================================================================
map_methods = [
    {"Container": "map/unordered_map", "Method": "operator[]", "Category": "Element Access", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const Key& key", "Arg_Explanation": "key: key to access", "Return_Type": "T&", "Description": "Access or insert element with key", "When_To_Use": "Access with default insert; simple syntax", "When_NOT_To_Use": "Don't want insertion (use find); const map", "Real_World_Freq": 10, "DSA_LeetCode_Freq": 10, "Example": "map[key] = value; int val = map[key];", "Notes": "Creates element if doesn't exist; can't use on const map", "Since_Version": "C++98", "Related": "at, insert, find"},
    {"Container": "map/unordered_map", "Method": "at", "Category": "Element Access", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const Key& key", "Arg_Explanation": "key: key to access", "Return_Type": "T&", "Description": "Access element (throws if not found)", "When_To_Use": "Safe access; no insertion; const maps", "When_NOT_To_Use": "Want default insert (use [])", "Real_World_Freq": 6, "DSA_LeetCode_Freq": 5, "Example": "int val = map.at(key);", "Notes": "Throws out_of_range; doesn't insert", "Since_Version": "C++11", "Related": "operator[], find"},
    {"Container": "map/unordered_map", "Method": "insert", "Category": "Modifiers", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const value_type& value", "Arg_Explanation": "value: pair<Key,Value> to insert", "Return_Type": "pair<iterator, bool>", "Description": "Inserts element if key doesn't exist", "When_To_Use": "Check if inserted; no overwrite", "When_NOT_To_Use": "Want to overwrite (use [] or insert_or_assign)", "Real_World_Freq": 8, "DSA_LeetCode_Freq": 8, "Example": "auto [it, inserted] = map.insert({key, value});", "Notes": "Returns pair: iterator and whether inserted", "Since_Version": "C++98", "Related": "operator[], emplace"},
    {"Container": "map/unordered_map", "Method": "emplace", "Category": "Modifiers", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "Args&&... args", "Arg_Explanation": "args: arguments to construct pair", "Return_Type": "pair<iterator, bool>", "Description": "Constructs element in-place", "When_To_Use": "Efficiency; avoid copy; complex values", "When_NOT_To_Use": "Simple insertion (insert or [] is clearer)", "Real_World_Freq": 7, "DSA_LeetCode_Freq": 6, "Example": "map.emplace(key, value);", "Notes": "More efficient than insert for complex types", "Since_Version": "C++11", "Related": "insert"},
    {"Container": "map/unordered_map", "Method": "erase", "Category": "Modifiers", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const Key& key or iterator pos", "Arg_Explanation": "key: key to remove or pos: iterator to element", "Return_Type": "size_t or iterator", "Description": "Removes element by key or iterator", "When_To_Use": "Remove elements", "When_NOT_To_Use": "Clearing all (use clear)", "Real_World_Freq": 8, "DSA_LeetCode_Freq": 9, "Example": "map.erase(key); map.erase(it);", "Notes": "By key returns count removed (0 or 1)", "Since_Version": "C++98", "Related": "clear"},
    {"Container": "map/unordered_map", "Method": "find", "Category": "Lookup", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const Key& key", "Arg_Explanation": "key: key to find", "Return_Type": "iterator", "Description": "Finds element by key", "When_To_Use": "Check existence; get iterator; no insertion", "When_NOT_To_Use": "Want default value (use [])", "Real_World_Freq": 9, "DSA_LeetCode_Freq": 10, "Example": "auto it = map.find(key); if (it != map.end()) {...}", "Notes": "Essential for checking existence; returns end() if not found", "Since_Version": "C++98", "Related": "count, contains"},
    {"Container": "map/unordered_map", "Method": "count", "Category": "Lookup", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const Key& key", "Arg_Explanation": "key: key to count", "Return_Type": "size_t", "Description": "Returns count of elements (0 or 1 for map)", "When_To_Use": "Check existence as bool", "When_NOT_To_Use": "Need iterator (use find); C++20 has contains", "Real_World_Freq": 7, "DSA_LeetCode_Freq": 8, "Example": "if (map.count(key)) {...}", "Notes": "Returns 0 or 1 for map (not multimap)", "Since_Version": "C++98", "Related": "find, contains"},
    {"Container": "map/unordered_map", "Method": "contains", "Category": "Lookup", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const Key& key", "Arg_Explanation": "key: key to check", "Return_Type": "bool", "Description": "Checks if key exists", "When_To_Use": "Simple existence check; modern C++", "When_NOT_To_Use": "Pre-C++20 (use count or find)", "Real_World_Freq": 8, "DSA_LeetCode_Freq": 9, "Example": "if (map.contains(key)) {...}", "Notes": "Added C++20; clearer than count", "Since_Version": "C++20", "Related": "count, find"},
    {"Container": "map/unordered_map", "Method": "size", "Category": "Capacity", "Time_Complexity": "O(1)", "Space_Complexity": "O(1)", "Arguments": "none", "Arg_Explanation": "", "Return_Type": "size_t", "Description": "Returns number of elements", "When_To_Use": "Get size; check emptiness", "When_NOT_To_Use": "Just checking empty (use empty())", "Real_World_Freq": 9, "DSA_LeetCode_Freq": 9, "Example": "size_t n = map.size();", "Notes": "Constant time", "Since_Version": "C++98", "Related": "empty"},
    {"Container": "map/unordered_map", "Method": "clear", "Category": "Modifiers", "Time_Complexity": "O(n)", "Space_Complexity": "O(1)", "Arguments": "none", "Arg_Explanation": "", "Return_Type": "void", "Description": "Removes all elements", "When_To_Use": "Empty map; reset state", "When_NOT_To_Use": "Removing few elements (use erase)", "Real_World_Freq": 7, "DSA_LeetCode_Freq": 7, "Example": "map.clear();", "Notes": "Size becomes 0", "Since_Version": "C++98", "Related": "erase"},
]

# ==============================================================================
# SET METHODS - std::set<T> and std::unordered_set<T>
# ==============================================================================
set_methods = [
    {"Container": "set/unordered_set", "Method": "insert", "Category": "Modifiers", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const T& value", "Arg_Explanation": "value: value to insert", "Return_Type": "pair<iterator, bool>", "Description": "Inserts element if not present", "When_To_Use": "Add unique elements", "When_NOT_To_Use": "Duplicates needed (use multiset)", "Real_World_Freq": 9, "DSA_LeetCode_Freq": 10, "Example": "auto [it, inserted] = set.insert(value);", "Notes": "Returns pair: iterator and whether inserted", "Since_Version": "C++98", "Related": "emplace, erase"},
    {"Container": "set/unordered_set", "Method": "erase", "Category": "Modifiers", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const T& value or iterator pos", "Arg_Explanation": "value: value to remove or pos: iterator", "Return_Type": "size_t or iterator", "Description": "Removes element", "When_To_Use": "Remove elements", "When_NOT_To_Use": "Clear all (use clear)", "Real_World_Freq": 8, "DSA_LeetCode_Freq": 9, "Example": "set.erase(value);", "Notes": "By value returns count removed", "Since_Version": "C++98", "Related": "insert, clear"},
    {"Container": "set/unordered_set", "Method": "find", "Category": "Lookup", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const T& value", "Arg_Explanation": "value: value to find", "Return_Type": "iterator", "Description": "Finds element", "When_To_Use": "Check existence with iterator", "When_NOT_To_Use": "Just bool (use count or contains)", "Real_World_Freq": 9, "DSA_LeetCode_Freq": 10, "Example": "auto it = set.find(value); if (it != set.end()) {...}", "Notes": "Returns end() if not found", "Since_Version": "C++98", "Related": "count, contains"},
    {"Container": "set/unordered_set", "Method": "count", "Category": "Lookup", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const T& value", "Arg_Explanation": "value: value to count", "Return_Type": "size_t", "Description": "Returns count (0 or 1 for set)", "When_To_Use": "Existence check as bool", "When_NOT_To_Use": "C++20 has contains; need iterator (use find)", "Real_World_Freq": 8, "DSA_LeetCode_Freq": 9, "Example": "if (set.count(value)) {...}", "Notes": "Returns 0 or 1", "Since_Version": "C++98", "Related": "find, contains"},
    {"Container": "set/unordered_set", "Method": "contains", "Category": "Lookup", "Time_Complexity": "O(log n) / O(1) avg", "Space_Complexity": "O(1)", "Arguments": "const T& value", "Arg_Explanation": "value: value to check", "Return_Type": "bool", "Description": "Checks if value exists", "When_To_Use": "Simple existence check; modern C++", "When_NOT_To_Use": "Pre-C++20 (use count or find)", "Real_World_Freq": 8, "DSA_LeetCode_Freq": 10, "Example": "if (set.contains(value)) {...}", "Notes": "Added C++20; very useful", "Since_Version": "C++20", "Related": "count, find"},
    {"Container": "set", "Method": "lower_bound", "Category": "Lookup", "Time_Complexity": "O(log n)", "Space_Complexity": "O(1)", "Arguments": "const T& value", "Arg_Explanation": "value: value to find", "Return_Type": "iterator", "Description": "First element >= value", "When_To_Use": "Range queries; ordered operations", "When_NOT_To_Use": "unordered_set (no ordering)", "Real_World_Freq": 6, "DSA_LeetCode_Freq": 8, "Example": "auto it = set.lower_bound(value);", "Notes": "Only for ordered set; very useful in DSA", "Since_Version": "C++98", "Related": "upper_bound, equal_range"},
    {"Container": "set", "Method": "upper_bound", "Category": "Lookup", "Time_Complexity": "O(log n)", "Space_Complexity": "O(1)", "Arguments": "const T& value", "Arg_Explanation": "value: value to find", "Return_Type": "iterator", "Description": "First element > value", "When_To_Use": "Range queries; ordered operations", "When_NOT_To_Use": "unordered_set (no ordering)", "Real_World_Freq": 5, "DSA_LeetCode_Freq": 8, "Example": "auto it = set.upper_bound(value);", "Notes": "Only for ordered set", "Since_Version": "C++98", "Related": "lower_bound"},
]

# ==============================================================================
# OTHER CONTAINERS - Quick reference
# ==============================================================================
other_containers = [
    {"Container": "deque", "Key_Methods": "push_front, pop_front, push_back, pop_back, operator[], at", "Use_Case": "Double-ended operations; both ends efficient", "vs_vector": "Better for front operations; slightly slower random access", "Real_World_Freq": 6, "DSA_LeetCode_Freq": 7},
    {"Container": "list", "Key_Methods": "push_front, push_back, insert, erase, splice, sort", "Use_Case": "Frequent insertions/deletions at known positions", "vs_vector": "No random access; O(1) insert/erase; poor cache locality", "Real_World_Freq": 4, "DSA_LeetCode_Freq": 3},
    {"Container": "forward_list", "Key_Methods": "push_front, insert_after, erase_after", "Use_Case": "Memory-constrained singly-linked list", "vs_list": "Half memory of list; no backward traversal", "Real_World_Freq": 2, "DSA_LeetCode_Freq": 2},
    {"Container": "stack", "Key_Methods": "push, pop, top, empty, size", "Use_Case": "LIFO operations; DFS, expression evaluation", "Notes": "Adapter (default: deque); no iterators", "Real_World_Freq": 8, "DSA_LeetCode_Freq": 9},
    {"Container": "queue", "Key_Methods": "push, pop, front, back, empty, size", "Use_Case": "FIFO operations; BFS, task queues", "Notes": "Adapter (default: deque); no iterators", "Real_World_Freq": 7, "DSA_LeetCode_Freq": 9},
    {"Container": "priority_queue", "Key_Methods": "push, pop, top, empty, size", "Use_Case": "Heap operations; Dijkstra, scheduling, top-k", "Notes": "Max heap by default; no iterators; adapter over vector", "Real_World_Freq": 7, "DSA_LeetCode_Freq": 10},
    {"Container": "array", "Key_Methods": "operator[], at, front, back, fill, size", "Use_Case": "Fixed-size array with STL interface", "vs_vector": "Fixed size; stack allocated; no dynamic growth", "Real_World_Freq": 6, "DSA_LeetCode_Freq": 5},
    {"Container": "multiset", "Key_Methods": "insert, erase, count, find, lower_bound, upper_bound", "Use_Case": "Sorted collection with duplicates", "vs_set": "Allows duplicates; count can be > 1", "Real_World_Freq": 4, "DSA_LeetCode_Freq": 6},
    {"Container": "multimap", "Key_Methods": "insert, erase, count, find, equal_range", "Use_Case": "Key-value with duplicate keys", "vs_map": "Multiple values per key", "Real_World_Freq": 3, "DSA_LeetCode_Freq": 4},
]


# ==============================================================================
# GENERATE EXCEL FILE
# ==============================================================================

print("Creating DataFrames...")

# Create DataFrames
df_algorithms = pd.DataFrame(stl_algorithms)
df_vector = pd.DataFrame(vector_methods)
df_map = pd.DataFrame(map_methods)
df_set = pd.DataFrame(set_methods)
df_other = pd.DataFrame(other_containers)

# Output file
output_file = Path("cpp_dsa_functions_catalog.xlsx")

print(f"Writing to {output_file}...")

# Write to Excel
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_algorithms.to_excel(writer, sheet_name='STL Algorithms', index=False)
    df_vector.to_excel(writer, sheet_name='Vector Methods', index=False)
    df_map.to_excel(writer, sheet_name='Map Methods', index=False)
    df_set.to_excel(writer, sheet_name='Set Methods', index=False)
    df_other.to_excel(writer, sheet_name='Other Containers', index=False)

# Load workbook for formatting
print("Applying formatting...")
wb = load_workbook(output_file)

# Format each sheet
colors = {
    'STL Algorithms': '4472C4',     # Blue
    'Vector Methods': '70AD47',      # Green
    'Map Methods': 'FFC000',         # Orange
    'Set Methods': '5B9BD5',         # Light Blue
    'Other Containers': 'A5A5A5'     # Gray
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    format_worksheet(ws, header_color=colors.get(sheet_name, '4472C4'))

# Save formatted workbook
wb.save(output_file)

print("\n" + "=" * 80)
print("✅ C++ DSA FUNCTIONS CATALOG GENERATED SUCCESSFULLY!")
print("=" * 80)
print(f"\n📁 File: {output_file.absolute()}")
print(f"\n📊 Sheets Created:")
print(f"   1. STL Algorithms     - {len(df_algorithms)} functions from <algorithm> and <numeric>")
print(f"   2. Vector Methods     - {len(df_vector)} methods for std::vector")
print(f"   3. Map Methods        - {len(df_map)} methods for std::map/unordered_map")
print(f"   4. Set Methods        - {len(df_set)} methods for std::set/unordered_set")
print(f"   5. Other Containers   - {len(df_other)} container summaries")
print(f"\n📈 Total entries: {len(df_algorithms) + len(df_vector) + len(df_map) + len(df_set) + len(df_other)}")
print("\n✨ Features:")
print("   • Color-coded headers for each sheet")
print("   • Auto-filtering enabled on all columns")
print("   • Frozen header row and first column")
print("   • Optimized column widths")
print("   • Alternating row colors for readability")
print("   • Real-world and DSA/LeetCode frequency ratings (1-10)")
print("   • Time and space complexity for all operations")
print("   • Comprehensive examples and usage notes")
print("   • C++ version information (C++98, C++11, C++17, C++20, C++23)")
print("\n🎯 Perfect for:")
print("   • Learning C++ STL algorithms and containers")
print("   • LeetCode and competitive programming preparation")
print("   • Quick reference during coding interviews")
print("   • Understanding time/space complexity")
print("   • Choosing the right algorithm/container for the job")
print("\n" + "=" * 80)

